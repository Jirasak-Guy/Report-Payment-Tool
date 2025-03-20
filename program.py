import tkinter as tk
from tkinter import filedialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import tkinter.font 
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import platform

# Global variables
input_path = ""
output_path = ""
new_output_folder = ""  # เพิ่มตัวแปรสำหรับเก็บ path โฟลเดอร์ปลายทาง

def shorten_path(path, max_length=40, max_filename_length=26):
    if len(path) <= max_length:
        return path
    path = os.path.normpath(path)
    directory, filename = os.path.split(path) # แยก path เป็นชื่อไฟล์กับ directory
    if len(filename) > max_filename_length:
        name, ext = os.path.splitext(filename)
        shortened_name = name[:max_filename_length - len(ext) - 3] + '...'
        filename = shortened_name + ext
    parts = directory.split(os.sep)  
    if len(parts) >= 3 :
        shortened_dir = os.sep.join([parts[0], "..."] + parts[-2:])
        return os.path.join(shortened_dir, filename)
    return os.path.join(directory, filename)

def browse_input_folder():
    global input_path
    folder_path = filedialog.askdirectory()
    if folder_path:
        input_path = folder_path
        input_label.config(text=shorten_path(folder_path))
        excel_files = [f for f in os.listdir(input_path) if f.endswith(('.xlsx', '.xls'))]
        status_label.config(text=f"พบ {len(excel_files)} ไฟล์ Excel")

def browse_output_folder():
    global output_path
    folder_path = filedialog.askdirectory()
    if folder_path:
        output_path = folder_path
        output_label.config(text=shorten_path(folder_path))

def open_output_folder():
    if new_output_folder:
        if platform.system() == "Windows":
            os.startfile(new_output_folder)
        elif platform.system() == "Darwin":  # macOS
            os.system(f"open {new_output_folder}")
        else:  # Linux
            os.system(f"xdg-open {new_output_folder}")

def process_files():
    global new_output_folder
    if not input_path or not output_path:
        status_label.config(text="กรุณาเลือกโฟลเดอร์ทั้งสอง")
        return
    
    excel_files = [f for f in os.listdir(input_path) if f.endswith(('.xlsx', '.xls'))]
    if not excel_files:
        status_label.config(text="ไม่พบไฟล์ Excel")
        return
    
    input_folder_name = os.path.basename(input_path)
    new_output_folder = os.path.join(output_path, f"{input_folder_name}_แก้ไขแล้ว")
    os.makedirs(new_output_folder, exist_ok=True)

    for file_name in excel_files:
        file_path = os.path.join(input_path, file_name)
        try:
            df = pd.read_excel(file_path)
            expected_columns = ['ลำดับ', 'รายการ', 'วันที่', 'ราคาต่อหน่วย', 'จำนวน', 'ราคาสุทธิ']
            if not all(col in df.columns for col in expected_columns):
                status_label.config(text=f"ไฟล์ {file_name} คอลัมน์ไม่ครบ")
                continue
            
            special_items_indices = []
            affected_bills_indices = set()
            new_order = 1
            current_bill_total = 0
            current_bill_start_idx = None

            for idx, row in df.iterrows():
                if isinstance(row['รายการ'], str) and row['รายการ'].startswith('ORR'):
                    if current_bill_start_idx is not None:
                        df.at[current_bill_start_idx, 'ราคาสุทธิ'] = round(current_bill_total)
                    df.at[idx, 'ลำดับ'] = new_order
                    new_order += 1
                    current_bill_total = 0
                    current_bill_start_idx = idx
                else:
                    df.at[idx, 'ลำดับ'] = ''
                    if isinstance(row['รายการ'], str) and '@' in row['รายการ']:
                        special_items_indices.append(idx)
                        if current_bill_start_idx is not None:
                            affected_bills_indices.add(current_bill_start_idx)
                        old_price = row['ราคาต่อหน่วย']
                        if pd.notna(old_price) and isinstance(old_price, (int, float)):
                            new_price = round((old_price - (old_price * 10 / 110)) * 1.03)
                            df.at[idx, 'ราคาต่อหน่วย'] = new_price
                            quantity = row['จำนวน']
                            if pd.notna(quantity) and isinstance(quantity, (int, float)):
                                new_net_price = new_price * quantity
                                df.at[idx, 'ราคาสุทธิ'] = new_net_price
                                current_bill_total += new_net_price
                    else:
                        net_price = row['ราคาสุทธิ']
                        if pd.notna(net_price) and isinstance(net_price, (int, float)):
                            current_bill_total += net_price

            if current_bill_start_idx is not None:
                df.at[current_bill_start_idx, 'ราคาสุทธิ'] = round(current_bill_total)

            output_file_name = f"{os.path.splitext(file_name)[0]}_แก้ไขแล้ว{os.path.splitext(file_name)[1]}"
            output_file = os.path.join(new_output_folder, output_file_name)
            df.to_excel(output_file, index=False)

            workbook = load_workbook(output_file)
            worksheet = workbook.active
            numeric_columns = ['ราคาต่อหน่วย', 'จำนวน', 'ราคาสุทธิ']
            col_indices = [df.columns.get_loc(col) + 1 for col in numeric_columns]
            price_per_unit_col = df.columns.get_loc('ราคาต่อหน่วย') + 1
            net_price_col = df.columns.get_loc('ราคาสุทธิ') + 1
            item_col = df.columns.get_loc('รายการ') + 1

            for col_idx in col_indices:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value is not None and isinstance(cell.value, (int, float)) and abs(cell.value) < 0.0001:
                        cell.number_format = '"-"'
                    else:
                        cell.number_format = '0'

            red_font = Font(color="FF0000")
            for idx in special_items_indices:
                row = idx + 2
                worksheet.cell(row=row, column=item_col).font = red_font
                worksheet.cell(row=row, column=price_per_unit_col).font = red_font
                worksheet.cell(row=row, column=net_price_col).font = red_font

            for idx in affected_bills_indices:
                worksheet.cell(row=idx + 2, column=net_price_col).font = red_font

            for i, column in enumerate(df.columns, 1):
                col_letter = chr(64 + i)
                max_length = max(df[column].astype(str).apply(len).max(), len(str(column)))
                if column in numeric_columns:
                    max_length = max(max_length, 1)
                worksheet.column_dimensions[col_letter].width = max_length

            workbook.save(output_file)
            status_label.config(text="ประมวลผลสำเร็จ" ,foreground="green")
            # แสดงปุ่ม "เปิดโฟลเดอร์" เมื่อประมวลผลสำเร็จ
            open_folder_button.pack(pady=10)

        except Exception as e:
            status_label.config(text=f"ข้อผิดพลาด: {file_name}",foreground="red")
            continue 

# GUI Setup
window = ttk.Window(themename="cosmo")
window.title("Report Payment Tool")
window.geometry("720x480")
window.resizable(False, False)
topic_btn = tkinter.font.Font( family = "TH Sarabun New",  size = 20,  weight = "bold") 
label_font = tkinter.font.Font( family = "TH Sarabun New",  size = 16) 

# Style configuration
style = ttk.Style()
style.configure("TButton", font=topic_btn, padding=5)
style.configure("TLabel", font=label_font)

# Main frame
main_frame = ttk.Frame(window, padding="20")
main_frame.pack(fill="both", expand=True)

# Input Folder Section
ttk.Label(main_frame, text="โฟลเดอร์ต้นทาง", font=topic_btn).pack(anchor="w", pady=(0, 5))
input_frame = ttk.Frame(main_frame)
input_frame.pack(fill="x", pady=(0, 10))
input_label = ttk.Label(input_frame, text="ยังไม่ได้เลือก", wraplength=450, foreground="#555555")
input_label.pack(side="left", fill="x", expand=True)
ttk.Button(input_frame, text="เลือก", command=browse_input_folder, bootstyle=PRIMARY).pack(side="right")


# Output Folder Section
ttk.Label(main_frame, text="โฟลเดอร์ปลายทาง", font=topic_btn).pack(anchor="w", pady=(0, 5))
output_frame = ttk.Frame(main_frame)
output_frame.pack(fill="x", pady=(0, 10))
output_label = ttk.Label(output_frame, text="ยังไม่ได้เลือก", wraplength=450, foreground="#555555")
output_label.pack(side="left", fill="x", expand=True)
ttk.Button(output_frame, text="เลือก", command=browse_output_folder, bootstyle=PRIMARY).pack(side="right")

# Status Label
status_label = ttk.Label(main_frame, text="พร้อมเริ่มต้น", foreground="#666666")
status_label.pack(pady=15)

# Process Button
ttk.Button(main_frame, text="ประมวลผล", command=process_files, width=20, bootstyle=PRIMARY).pack()

open_folder_button = ttk.Button(main_frame, text="เปิดโฟลเดอร์", command=open_output_folder, width=20, bootstyle=SUCCESS)

# Center window
window.update_idletasks()
width, height = window.winfo_width(), window.winfo_height()
x = (window.winfo_screenwidth() // 2) - (width // 2)
y = (window.winfo_screenheight() // 2) - (height // 2)
window.geometry(f"+{x}+{y}")

window.mainloop()